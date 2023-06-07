import pandas as pd
import json
import openpyxl p[rasd]
import datetime
from datetime import time
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import re
import os 

fname = '/Users/669966_Standard/Library/CloudStorage/OneDrive-Amtrak/Python_CDHK_SDHM_Parse/hello/CDHK_SDHM_B2B.xlsx'
al = Alignment(horizontal="left", vertical="top", wrap_text=True)
hl = Alignment(horizontal="left", vertical="top")
redFill = PatternFill(start_color='00C0C0C0',
                      end_color='00C0C0C0', fill_type='solid')

redFillColumn2 = PatternFill(start_color='00CCCCFF',
                             end_color='00CCCCFF', fill_type='solid')

redFillColumn3 = PatternFill(start_color='00808000',
                             end_color='00808000', fill_type='solid')
redFillColumn4 = PatternFill(start_color='0099CCFF',
                             end_color='0099CCFF', fill_type='solid')

# load excel with its path
wrkbk = openpyxl.load_workbook(fname)
sh = wrkbk.active
a_true_alias = False
# iterate through excel and display data
for i in range(1, sh.max_row+1):
    if (a_true_alias):
        break
    # print("\n")
    print("Row ", i, " data :")
    
    cell_obj = sh.cell(row=i, column=6)
    if cell_obj.value is None:
        a_true_alias = True
        break
    
    for z in range(1, sh.max_column+1):
        # print("Row " + str(i) + "  and column  :"+str(j))
        if ( z >= 6 ):
            continue
        # if i > 1:
        if z == 2:
            cell_obj = sh.cell(row=i, column=z)
            
            try:
                y = json.loads(cell_obj.value)
            except:
                if cell_obj.value is not None:
                    print("No JSON present for GLOBAL KINESIS at "+str(cell_obj.coordinate))
                else :
                    print("BLANK VALUE PRESENT for GLOBAL KINESIS at "+str(cell_obj.coordinate))
                continue

            
            # ticket section testing
            try:
                my_string = ""
                ticketNumber = ""
                for j in range(0, len(y['PNRpublishPB']['TicketSection']['TicketItem'])):
                    if j != 0:
                        if (len(ticketNumber) == 0):
                            ticketNumber = y['PNRpublishPB']['TicketSection']['TicketItem'][j]['TicketNumber']
                        else:
                            ticketNumber = ticketNumber+"," + \
                                y['PNRpublishPB']['TicketSection']['TicketItem'][j]['TicketNumber']
            except:
                ticketNumber = "Ticket number is not present"
            # ticket section testing



            segmenntInformation = ""
            for j in range(0, len(y['PNRpublishPB']['SegmentSection']['Segment'])):
                if (len(segmenntInformation) == 0):
                    segmenntInformation = "\nOrigin/Destination Station for Segment "+str(j+1)+"-->>"+y['PNRpublishPB']['SegmentSection']['Segment'][j]['OriginLocationCode']+'/'+y['PNRpublishPB']['SegmentSection']['Segment'][j]['DestinationCode']+"\nDeparture and Arrival Date/Time for Segment "+str(
                        j+1)+"-->>"+y['PNRpublishPB']['SegmentSection']['Segment'][j]['DepartureDateTime']+'/'+y['PNRpublishPB']['SegmentSection']['Segment'][j]['ArrivalDateTime']+'\n'
                else:
                    segmenntInformation = segmenntInformation+"\n"+"Origin/Destination Station for Segment "+str(j+1)+"-->>"+y['PNRpublishPB']['SegmentSection']['Segment'][j]['OriginLocationCode']+'/'+y['PNRpublishPB']['SegmentSection']['Segment'][j][
                        'DestinationCode']+"\nDeparture and Arrival Date/Time for Segment "+str(j+1)+"-->>"+y['PNRpublishPB']['SegmentSection']['Segment'][j]['DepartureDateTime']+'/'+y['PNRpublishPB']['SegmentSection']['Segment'][j]['ArrivalDateTime']+'\n'

            address = ""
            for j in range(0, len(y['PNRpublishPB']['FiveFieldSection']['Miscellaneous'])):
                if (y['PNRpublishPB']['FiveFieldSection']['Miscellaneous'][j]['MiscTenCode'] == 'BAI'):
                    address = y['PNRpublishPB']['FiveFieldSection']['Miscellaneous'][j]['MiscTenText']

            passengerName = ""
            for j in range(0, len(y['PNRpublishPB']['PassengerSection']['Passenger'])):
                if (len(passengerName) == 0):
                    passengerName = "Passenger Name " + \
                        str(j+1)+"-->>"+y['PNRpublishPB']['PassengerSection']['Passenger'][j]['FirstName'] + \
                        ' ' + \
                        y['PNRpublishPB']['PassengerSection']['Passenger'][j]['LastName']+'\n'
                else:
                    passengerName = passengerName+"Passenger Name " + \
                        str(j+1)+"-->>"+y['PNRpublishPB']['PassengerSection']['Passenger'][j]['FirstName'] + \
                        ' ' + \
                        y['PNRpublishPB']['PassengerSection']['Passenger'][j]['LastName']+'\n'

            agrNumber = ""
            try:
                agrNumber = y['PNRpublishPB']['PassengerSection']['Passenger'][0]['AGRloyaltyNbr']
            except:
                agrNumber = "Booking  Done as guest"

            costOfTicket = ""
            try:
                costOfTicket = y['PNRpublishPB']['TicketReceiptSection']['TotalTktAmt']
            except:
                costOfTicket = "JSON modified due to size "

            my_string = "PNR-->>"+y['PNRpublishPB']['RCN']+'\n'+passengerName+'\n'+"emailAddress-->>"+y['PNRpublishPB']['PassengerSection']['EmailAddr1']+'\n' + \
                "numberofSegment-->>"+y['PNRpublishPB']['SegmentSection']['NumberOfSegments']+'\n'+"numberOfPassenger-->>" +y['PNRpublishPB']['PassengerSection']['NbrOfPassengers']+'\n'+ \
                '\n'+segmenntInformation+'\n' +  \
                "AGR Number-->>"+agrNumber+'\n' +  \
                "bookingTime-->>"+y['PNRpublishPB']['currentDatetime']+'\n'+"bookingTime-->>"+y['PNRpublishPB']['PNRCreateDatetime'] + \
                "DOB-->>"+y['PNRpublishPB']['PassengerSection']['Passenger'][0]['PassengerBirthdate']+'\n'+"phoneNumber-->>"+  \
                y['PNRpublishPB']['PhoneSection']['PassengerContact'][0]['AreaCode'] + \
                y['PNRpublishPB']['PhoneSection']['PassengerContact'][0]['PhoneNbr']+'\n'+"costOfTicket-->>"+costOfTicket + \
                '\n'+"completeCustomerAddress-->>" + \
                address + \
                '\n\n'+"Ticket Number "+ticketNumber+'\n'
            # "Passenger Name-->>"+y['PNRpublishPB']['PassengerSection']['Passenger'][0]['FirstName']+" "+y['PNRpublishPB']['PassengerSection']['Passenger'][0]['LastName']
            sh['G'+re.search(r"\d+(\.\d+)?", cell_obj.coordinate).group(0)] = my_string
            sh['G'+re.search(r"\d+(\.\d+)?", cell_obj.coordinate).group(0)].font = Font(bold=True)
            sh['G'+re.search(r"\d+(\.\d+)?", cell_obj.coordinate).group(0)].alignment = al
            sh['G'+re.search(r"\d+(\.\d+)?", cell_obj.coordinate).group(0)].fill = redFill
            sh.column_dimensions['F'].width = 80

        if z == 3:
            cell_obj = sh.cell(row=i, column=z)
            try:
                try:
                    y = json.loads(cell_obj.value)
                except:
                    if cell_obj.value is not None:
                        print("No JSON present for PNR PROCESSOR LAMBDA at "+str(cell_obj.coordinate))
                    else :
                        print("BLANK VALUE PRESENT for PNR PROCESSOR LAMBDA at "+str(cell_obj.coordinate))
                    continue

                if cell_obj.value is not None:
                    y = json.loads(cell_obj.value)
                    my_string = ""

                    pnrNumber = ""
                    if isinstance(y['custPnr']['pnrNumber'], list):
                        test_list = list(set(y['custPnr']['pnrNumber']))
                        for j in range(0, len(test_list)):
                            if (len(pnrNumber) == 0):
                                pnrNumber = test_list[j]
                            else:
                                pnrNumber = pnrNumber+"," + test_list[j]
                    else:
                        pnrNumber = y['custPnr']['pnrNumber']

                    ticketNumber = ""
                    if isinstance(y['custPnr']['ticketNumber'], list):
                        test_list = []
                        test_list = list(set(y['custPnr']['ticketNumber']))
                        for j in range(0, len(test_list)):
                            if (len(ticketNumber) == 0):
                                ticketNumber = test_list[j]
                            else:
                                ticketNumber = ticketNumber+"," + test_list[j]
                    else:
                        ticketNumber = y['custPnr']['ticketNumber']

                    sourceSystemName = ""
                    if isinstance(y['sourceSystemId'], list):
                        test_list = []
                        test_list = list(set(y['sourceSystemId']))
                        for j in range(0, len(test_list)):
                            if (len(sourceSystemName) == 0):
                                sourceSystemName = test_list[j]
                            else:
                                sourceSystemName = sourceSystemName + \
                                    "," + test_list[j]
                    else:
                        sourceSystemName = y['sourceSystemId']

                    channelId = ""
                    if isinstance(y['channelId'], list):
                        test_list = []
                        test_list = list(set(y['channelId']))
                        for j in range(0, len(test_list)):
                            if (len(channelId) == 0):
                                channelId = test_list[j]
                            else:
                                channelId = channelId+"," + test_list[j]
                    else:
                        channelId = y['channelId']

                    pnrCreateDate = ""
                    if isinstance(y['custPnr']['pnrCreateDate'], list):
                        test_list = []
                        test_list = list(set(y['custPnr']['pnrCreateDate']))
                        for j in range(0, len(test_list)):
                            if (len(pnrCreateDate) == 0):
                                pnrCreateDate = test_list[j]
                            else:
                                pnrCreateDate = pnrCreateDate + \
                                    "," + test_list[j]
                    else:
                        pnrCreateDate = y['custPnr']['pnrCreateDate']

                    passennngerName = ""
                    if isinstance(y['customerProfile']['firstName'], list):
                        test_list = []
                        if (len(y['customerProfile']['firstName']) == len(y['customerProfile']['lastName'])):
                            for j in range(0, len(y['customerProfile']['firstName'])):
                                if (len(passennngerName) == 0):
                                    passennngerName = y['customerProfile']['firstName'][j] + \
                                        " "+y['customerProfile']['lastName'][j]
                                else:
                                    passennngerName = passennngerName+"," + \
                                        y['customerProfile']['firstName'][j]+" " + \
                                        y['customerProfile']['lastName'][j]
                    else:
                        passennngerName = y['customerProfile']['firstName'] + \
                            " "+y['customerProfile']['lastName']

                    emailAddress = ""
                    test_list = []
                    for j in range(0, len(y['customerProfile']['custEmail'])):
                        if (len(emailAddress) == 0):
                            emailAddress = y['customerProfile']['custEmail'][j]['emailAddress']
                        else:
                            emailAddress = emailAddress +","+ y['customerProfile']['custEmail'][j]['emailAddress']

                    agrNumber = ""
                    try:
                        for j in range(0, len(y['customerProfile']['custAgr']['agrNumber'])):
                            if (len(agrNumber) == 0):
                                agrNumber = y['customerProfile']['custAgr']['agrNumber'][j]
                            else:
                                agrNumber = agrNumber   +","+ y['customerProfile']['custAgr']['agrNumber'][j]
                    except:
                        agrNumber="Booking done as guest"
                    
                    custAddress = ""
                    try:
                        custAddress = y['customerProfile']['custAddress'][0]['line1Address']+" "+y['customerProfile']['custAddress'][0]['line2Address']+" " + \
                        y['customerProfile']['custAddress'][0]['cityName']+" "+y['customerProfile']['custAddress'][0]['countryCode']+" " + \
                        y['customerProfile']['custAddress'][0]['postalBaseCode']
                    except:
                        custAddress="Address is not present as payment was not done from Credit Card"

                    customerPhone = ""
                    for j in range(0, len(y['customerProfile']['custPhone'])):
                        if (len(customerPhone) == 0):
                            customerPhone = y['customerProfile']['custPhone'][j]['phoneNumber']
                        else:
                            customerPhone = customerPhone  +","+  y['customerProfile']['custPhone'][j]['phoneNumber']

                    my_string = "sourceSystemName-->>"+sourceSystemName + \
                        "\nchannelId-->>"+channelId+'\n'+"Reservation-->>"+pnrNumber + \
                        "\npnrCreateDate-->>"+pnrCreateDate+'\n'+"ticketNumber-->>"+ticketNumber + \
                        "\nPassenger -->>"+passennngerName + \
                        "\nAGR Number -->>"+agrNumber + \
                        "\nEmail Address-->>"+emailAddress+'\n'+"Customer Address-->>"+custAddress + \
                        "\nCustomer Phone-->>"+customerPhone

                    sh['H'+re.search(r"\d+(\.\d+)?", cell_obj.coordinate).group(0)] = my_string
                    sh['H'+re.search(r"\d+(\.\d+)?", cell_obj.coordinate).group(0)].font = Font(bold=True)
                    sh['H'+re.search(r"\d+(\.\d+)?", cell_obj.coordinate).group(0)].alignment = al
                    sh['H'+re.search(r"\d+(\.\d+)?", cell_obj.coordinate).group(0)].fill = redFillColumn2
                    sh.column_dimensions['G'].width = 80
            except ValueError as e:
                continue

        if z == 4:
            cell_obj = sh.cell(row=i, column=z)
            print("value of cell position is "+str(cell_obj.coordinate))
            try:
                try:
                    y = json.loads(cell_obj.value)
                except:
                    if cell_obj.value is not None:
                        print("No JSON present for Dynnamo DB at "+str(cell_obj.coordinate))
                    else :
                        print("BLANK VALUE PRESENT for Dynnamo DB at "+str(cell_obj.coordinate))
                    continue

                if cell_obj.value is not None:
                    y = json.loads(cell_obj.value)
                    my_string = ""
                    pnrNumber = ""
                    if isinstance(y['pnr']['S'], list):
                        test_list = list(set(y['pnr']['S']))
                        for j in range(0, len(test_list)):
                            if (len(pnrNumber) == 0):
                                pnrNumber = test_list[j]
                            else:
                                pnrNumber = pnrNumber+"," + test_list[j]
                    else:
                        pnrNumber = y['pnr']['S']

                    ticketNumber = ""
                    if isinstance(y['tn']['S'], list):
                        test_list = list(set(y['tn']['S']))
                        for j in range(0, len(test_list)):
                            if (len(ticketNumber) == 0):
                                ticketNumber = test_list[j]
                            else:
                                ticketNumber = ticketNumber+"," + test_list[j]
                    else:
                        ticketNumber = y['tn']['S']

                    agrNumber = ""
                    try:
                        if isinstance(y['ano']['S'], list):
                            test_list = list(set(y['ano']['S']))
                            for j in range(0, len(test_list)):
                                if (len(agrNumber) == 0):
                                    agrNumber = test_list[j]
                                else:
                                    agrNumber = agrNumber+"," + test_list[j]
                        else:
                            agrNumber = y['ano']['S']
                    except:
                        agrNumber="Booking done as guest"
                    
                    address = ""
                    try:
                        if isinstance(y['add']['M']['al1']['S'], list):
                            address = str(y['add']['M']['al1']['S'][0])+','+str(y['add']['M']['al2']['S'][0]) + ", " + str(
                            y['add']['M']['ct']['S'][0])+"," + str(y['add']['M']['cy']['S'][0]) + ","+str(y['add']['M']['pc']['S'][0])
                        else:
                            address = str(y['add']['M']['al1']['S'])+','+str(y['add']['M']['al2']['S']) + ", " + str(
                            y['add']['M']['ct']['S'])+"," + str(y['add']['M']['cy']['S']) + ","+str(y['add']['M']['pc']['S'])
                    except:
                        address="Address is not present as payment was not done from Credit Card"

                    emailAddress = ""
                    if isinstance(y['em']['S'], list):
                        for j in range(0, len(y['em']['S'])):
                            if (len(emailAddress) == 0):
                                emailAddress = y['em']['S'][j]
                            else:
                                emailAddress = emailAddress+"," + y['em']['S'][j]
                    else:
                        emailAddress = str(y['em']['S'])

                    # testing
                    name = ""
                    if isinstance(y['fn']['S'], list):
                        for j in range(0, len(y['fn']['S'])):
                            if (len(name) == 0):
                                name = y['fn']['S'][j]+" "+y['ln']['S'][j]
                            else:
                                name = name+"," + \
                                    y['fn']['S'][j]+" "+y['ln']['S'][j]
                    else:
                        name = y['fn']['S']+" "+y['ln']['S']
                    # testing
                    # testing
                    phoneNumber = ""
                    if isinstance(y['pph']['L'], list):
                        for j in range(0, len(y['pph']['L'])):
                            if (len(phoneNumber) == 0):
                                phoneNumber = str(
                                    y['pph']['L'][j]['M']['accd']['S'])+""+str(y['pph']['L'][j]['M']['phno']['S'])
                            else:
                                phoneNumber = phoneNumber+"," + \
                                    str(y['pph']['L'][j]['M']['accd']['S']) + \
                                    ""+str(y['pph']['L'][j]['M']['phno']['S'])
                    else:
                        phoneNumber = str(
                            y['pph']['L'][j]['M']['accd']['S'])+""+str(y['pph']['L'][j]['M']['phno']['S'])
                    # testing

                    my_string = ""
                    my_string = "PNR NUMBER-->>"+pnrNumber + \
                        "\nAGR Number-->>"+agrNumber + \
                        "\nTicket NUMBER-->>"+ticketNumber+'\n'+"Address-->>"+address + \
                        "\nEmail Address-->>"+emailAddress+'\n'+"Name-->>"+name + \
                        "\nPhone Number -->>"+phoneNumber

                    sh['I'+re.search(r"\d+(\.\d+)?", cell_obj.coordinate).group(0)] = my_string
                    sh['I'+re.search(r"\d+(\.\d+)?", cell_obj.coordinate).group(0)].font = Font(bold=True)
                    sh['I'+re.search(r"\d+(\.\d+)?", cell_obj.coordinate).group(0)].alignment = al
                    sh['I'+re.search(r"\d+(\.\d+)?", cell_obj.coordinate).group(0)].fill = redFillColumn3
                    sh.column_dimensions['H'].width = 60
            # testing
                
            # testing
            except ValueError as e:
                continue

        if z == 5:
            cell_obj = sh.cell(row=i, column=z)
            print("value of cell position is "+str(cell_obj.coordinate))
            try:
                try:
                    y = json.loads(cell_obj.value)
                except:
                    if cell_obj.value is not None:
                        print("No JSON present for CDH payload at "+str(cell_obj.coordinate))
                    else :
                        print("BLANK VALUE PRESENT for CDH payload at "+str(cell_obj.coordinate))
                    continue

                if cell_obj.value is not None:
                    y = json.loads(cell_obj.value)
                    my_string = ""
                    # testing
                    pnrNumber = ""
                    if isinstance(y['PNR_NBR'], list):
                        test_list = list(set(y['PNR_NBR']))
                        for j in range(0, len(test_list)):
                            if (len(pnrNumber) == 0):
                                pnrNumber = test_list[j]
                            else:
                                pnrNumber = pnrNumber+"," + test_list[j]
                    else:
                        pnrNumber = y['PNR_NBR']

                    ticketNumber = ""
                    if isinstance(y['TCKT_NBR'], list):
                        test_list = list(set(y['TCKT_NBR']))
                        for j in range(0, len(test_list)):
                            if (len(ticketNumber) == 0):
                                ticketNumber = test_list[j]
                            else:
                                ticketNumber = ticketNumber+"," + test_list[j]
                    else:
                        ticketNumber = y['TCKT_NBR']

                    pnrCreateDate = ""
                    if isinstance(y['PNR_CREATE_DT'], list):
                        test_list = list(set(y['PNR_CREATE_DT']))
                        for j in range(0, len(test_list)):
                            if (len(pnrCreateDate) == 0):
                                pnrCreateDate = test_list[j]
                            else:
                                pnrCreateDate = pnrCreateDate + \
                                    "," + test_list[j]
                    else:
                        pnrCreateDate = y['PNR_CREATE_DT']

                    createDate = ""
                    if isinstance(y['CREATE_DATE'], list):
                        test_list = list(set(y['CREATE_DATE']))
                        for j in range(0, len(test_list)):
                            if (len(createDate) == 0):
                                createDate = test_list[j]
                            else:
                                createDate = createDate+"," + test_list[j]
                    else:
                        createDate = y['CREATE_DATE']

                    passengerName = ""
                    if isinstance(y['NAME'], list):
                        test_list = list(set(y['NAME']))
                        for j in range(0, len(test_list)):
                            if (len(passengerName) == 0):
                                passengerName = test_list[j]
                            else:
                                passengerName = passengerName + \
                                    "," + test_list[j]
                    else:
                        passengerName = y['NAME']

                    emailAddress = ""
                    if isinstance(y['EMAIL_ADDR'], list):
                        test_list = list(set(y['EMAIL_ADDR']))
                        for j in range(0, len(test_list)):
                            if (len(emailAddress) == 0):
                                if test_list[j] is not None:
                                    emailAddress = test_list[j]
                            else:
                                if test_list[j] is not None:
                                    emailAddress = emailAddress+"," + test_list[j]
                    else:
                        emailAddress = y['EMAIL_ADDR']

                    phoneNumber = ""
                    if isinstance(y['PHN_NBR'], list):
                        test_list = list(set(y['PHN_NBR']))
                        for j in range(0, len(test_list)):
                            if (len(phoneNumber) == 0):
                                phoneNumber = test_list[j]
                            else:
                                phoneNumber = phoneNumber+"," + test_list[j]
                    else:
                        phoneNumber = y['PHN_NBR']

                    address = ""
                    if isinstance(y['ADDRESS'], list):
                        test_list = list(set(y['ADDRESS']))
                        for j in range(0, len(test_list)):
                            if (len(address) == 0):
                                if test_list[j] is not None:
                                    address = test_list[j]
                            else:
                                if test_list[j] is not None:
                                    address = address+"," + test_list[j]
                    else:
                        address = y['ADDRESS']

                    my_string = ""
                    my_string = "PNR NUMBER-->>"+pnrNumber + \
                        "\nTicket NUMBER-->>"+ticketNumber+'\n'+"PNR Create Date-->>"+pnrCreateDate + \
                        "\nCDH Create Date-->>"+createDate+'\n'+"Passenger Name-->>"+passengerName + \
                        "\nEmail Address-->>"+emailAddress+'\n'+"Phone Number-->>"+str(phoneNumber) + \
                        "\nAddress -->>"+address+'\n'

                    sh['J'+re.search(r"\d+(\.\d+)?", cell_obj.coordinate).group(0)] = my_string
                    sh['J'+re.search(r"\d+(\.\d+)?", cell_obj.coordinate).group(0)].font = Font(bold=True)
                    sh['J'+re.search(r"\d+(\.\d+)?", cell_obj.coordinate).group(0)].alignment = al
                    sh['J'+re.search(r"\d+(\.\d+)?", cell_obj.coordinate).group(0)].fill = redFillColumn4
                    sh.column_dimensions['I'].width = 60
                    sh.column_dimensions['J'].width = 70

            # testing
            except ValueError as e:
                continue
# print(x)


sh.column_dimensions['B'].width = 50
sh.column_dimensions['C'].width = 40
sh.column_dimensions['D'].width = 33
sh.column_dimensions['F'].width = 29
sh.column_dimensions['A'].width = 13.5
sh['B1'].font = Font(bold=True, size=15)
sh['B1'].alignment = hl
sh['B1'].fill = redFill

sh['C1'].font = Font(bold=True, size=15)
sh['C1'].alignment = hl
sh['C1'].fill = redFill

sh['D1'].font = Font(bold=True, size=15)
sh['D1'].alignment = hl
sh['D1'].fill = redFill

sh['E1'].font = Font(bold=True, size=15)
sh['E1'].alignment = hl
sh['E1'].fill = redFill

sh['F1'].font = Font(bold=True, size=15)
sh['F1'].alignment = hl
sh['F1'].fill = redFill

sh['G1'] = 'Parsed Global Kinesis Stream'
sh['H1'] = 'Parsed pnr-processor-lambda'
sh['I1'] = 'Parsed pnr-records Dynamo DB'
sh['J1'] = 'Parsed CDH Record'

sh['G1'].font = Font(bold=True, size=15)
sh['G1'].alignment = hl
sh['G1'].fill = redFill

sh['H1'].font = Font(bold=True, size=15)
sh['H1'].alignment = hl
sh['H1'].fill = redFill

sh['I1'].font = Font(bold=True, size=15)
sh['I1'].alignment = hl
sh['I1'].fill = redFill

sh['J1'].font = Font(bold=True, size=15)
sh['J1'].alignment = hl
sh['J1'].fill = redFill

x = datetime.datetime.now()
saveFileName=""
saveFileName=str(x).replace(":", "_")
saveFileName=str(saveFileName).replace(".", "_")
saveFileName=str(saveFileName).replace("-", "_")
fileName = "/Users/669966_Standard/Library/CloudStorage/OneDrive-Amtrak/Python_CDHK_SDHM_Parse/hello/Parsed_Json/CDHK_SDHM_B2B_"+saveFileName+".xlsx"
wrkbk.save(fileName)
os.system("open -a 'Microsoft Excel.app' '%s'" % fileName)
print("File saved and opened at "+fileName)